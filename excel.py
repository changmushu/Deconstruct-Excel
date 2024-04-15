from openpyxl import load_workbook

def generate_validation_code(file_path):
    # 加载Excel文件
    wb = load_workbook(file_path)
    sheet = wb.active
    
    # 读取数据并生成代码
    code_lines = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # print("Row:", row)  # 添加这行进行调试
        field_name, is_required, api_field = row
        is_required = 'true' if is_required == '是' else 'false'
        validation_line = f"{api_field}: [{{ required: {is_required.lower()}, message: '请输入{field_name}', trigger: 'change' }}],"
        code_lines.append(validation_line)
    
    # 拼接代码
    code = "\n".join(code_lines)

        # 将结果写入新的文本文件
    with open(output_text_file, "w") as f:
        f.write(code)
    
    return code

# 用法示例
excel_file_path = "input.xlsx"
output_text_file = "output_code.txt"
generated_code = generate_validation_code(excel_file_path)
print(generated_code)
